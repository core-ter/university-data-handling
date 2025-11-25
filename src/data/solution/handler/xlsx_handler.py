import os
from typing import List
from datetime import date
import openpyxl
from openpyxl import Workbook

from data.solution.model import Department, Project

def write_departments(departments: List[Department], path: str, file_name: str = "departments.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Departments"
    
    ws.append(["id", "name", "floor"])
    for dept in departments:
        ws.append([dept.id, dept.name, dept.floor])
        
    wb.save(os.path.join(path, file_name))

def read_departments(path: str, file_name: str = "departments.xlsx") -> List[Department]:
    wb = openpyxl.load_workbook(os.path.join(path, file_name))
    ws = wb["Departments"]
    
    departments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        departments.append(Department(row[0], row[1], int(row[2])))
        
    return departments

def write_projects(projects: List[Project], path: str, file_name: str = "projects.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    
    ws.append(["id", "name", "budget", "deadline", "status"])
    for proj in projects:
        ws.append([proj.id, proj.name, proj.budget, proj.deadline, proj.status])
        
    wb.save(os.path.join(path, file_name))

def read_projects(path: str, file_name: str = "projects.xlsx") -> List[Project]:
    wb = openpyxl.load_workbook(os.path.join(path, file_name))
    ws = wb["Projects"]
    
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        # Excel might read date as datetime or string
        deadline = row[3]
        if hasattr(deadline, 'date'):
            deadline = deadline.date()
        elif isinstance(deadline, str):
            deadline = date.fromisoformat(deadline)
            
        projects.append(Project(row[0], row[1], int(row[2]), deadline, row[4]))
        
    return projects
